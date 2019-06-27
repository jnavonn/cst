#!/bin/python
# -*- coding: utf-8 -*-

import wx
import os
import datetime
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment

class MyFrame(wx.Frame):

    def search_identifiant_action(self):
        LstAction = []
        query = "SELECT distinct(TK) FROM equities ORDER BY TK"

        # **************************************************************************************
        # Récupérer la liste des actions dans la base de données MySQL
        # **************************************************************************************
        conn = mysql.connector.connect(host="35.158.172.159",
                                       user="cst", password="cst-GroupeHN-2019",
                                       database="sedec")

        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        for row in rows:
            LstAction.append(row[0])

        cursor = conn.cursor()

        return LstAction

    
    def search_price_by_action(self, identifiant_action):
        query = "SELECT Datec, TRI, PX FROM prices where TK = %(TK)s"

        # **************************************************************************************
        # Récupérer tous les enregistrements de l'action sélectionnée dans la base de données MySQL
        # **************************************************************************************
        conn = mysql.connector.connect(host="35.158.172.159",
                                       user="cst", password="cst-GroupeHN-2019",
                                       database="sedec")

        cursor = conn.cursor()
        cursor.execute(query, {'TK': identifiant_action})
        rows = cursor.fetchall()

        cursor = conn.cursor()

        return rows

    def write_in_file(self, identifiant_action, rows):

        # **************************************************************************************
        # Ecrire du résultat dans un fichier Excel
        # **************************************************************************************
        # Créer un classeur
        wb = Workbook()
        # Créer une feuille
        ws = wb.active
        ws.title = identifiant_action

        # Creér un style nommé pour les titres de colonne dans le fichier
        StyleLabel = NamedStyle(name="StyleLabel")
        StyleLabel.font = Font(bold=True, color="FFFFFF")
        StyleLabel.alignment = Alignment(horizontal='center', vertical='center')
        StyleLabel.fill = PatternFill(fgColor="318CE7", fill_type="solid")
        StyleLabel.border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   top=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))

        # Creér un style nommé pour les valeurs de colonne dans le fichier
        StyleValue = NamedStyle(name="StyleValue")
        StyleValue.alignment = Alignment(horizontal='center', vertical='center')
        StyleValue.border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'))

        # Définir la largeur des colonnes du fichier
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Ecrire les titres de colonnes dans le fichier
        ws.cell(row=1, column=1, value="Date").style = StyleLabel
        ws.cell(row=1, column=2, value="TRI").style = StyleLabel
        ws.cell(row=1, column=3, value="PX").style = StyleLabel

        # Remplir le fichier avec toutes les valeurs des colonnes
        i = 2
        j = 0
        for row in rows:
            date_time_obj = datetime.datetime.strptime(str(row[0]), '%Y-%m-%d')
            ws.cell(row=i, column=1, value=date_time_obj.date().strftime('%d/%m/%Y')).style = StyleValue
            ws.cell(row=i, column=2, value=row[1]).style = StyleValue
            ws.cell(row=i, column=3, value=row[2]).style = StyleValue
            if j == 1:
                ws.cell(row=i, column=1).fill = PatternFill(fgColor="DFF2FF", fill_type="solid")
                ws.cell(row=i, column=2).fill = PatternFill(fgColor="DFF2FF", fill_type="solid")
                ws.cell(row=i, column=3).fill = PatternFill(fgColor="DFF2FF", fill_type="solid")
                j = 0
            else:
                j = 1
            i = i + 1

        # Mettre une bordure inférieure sur la dernière ligne
        ws.cell(row=i-1, column=1).border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))
        ws.cell(row=i-1, column=2).border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))
        ws.cell(row=i-1, column=3).border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))

        wb.save(identifiant_action+".xlsx")

        return True

    def __init__(self, titre):
        wx.Frame.__init__(self, None, 1, title=titre)

        # Définir les différentes zones dans la fenêtre
        frameSizer = wx.BoxSizer(wx.HORIZONTAL)
        panelSizer = wx.BoxSizer(wx.VERTICAL)
        conteneur = wx.Panel(self, 1)

        # Ajouter un texte dans le conteneur
        textAction = wx.StaticText(conteneur, 1, "Liste des actions boursières", style=wx.ALIGN_CENTRE)
        textbtnEcrire = wx.StaticText(conteneur, 1, "Liste des prix de l'action par date", style=wx.ALIGN_CENTRE)

        # Ajouter une liste deroulante dans le conteneur
        self.lstAction = wx.Choice(conteneur)

        # Ajouter un bouton dans le conteneur
        self.btnEcrire = wx.Button(conteneur, -1, "Ecrire")

        # Ajouter verticalement tous les contrôles
        panelSizer.AddSpacer(20)
        panelSizer.Add(textAction, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.lstAction, 1, wx.ALIGN_CENTRE)
        panelSizer.AddSpacer(20)
        panelSizer.Add(textbtnEcrire, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.btnEcrire, 1, wx.ALIGN_CENTRE)
        panelSizer.AddSpacer(20)

        conteneur.SetSizer(panelSizer)
        frameSizer.Add(conteneur, 1, wx.EXPAND)
        self.SetSizer(frameSizer)
        frameSizer.SetSizeHints(self)
        self.SetSize((400, 230))

        # Remplisssage de la liste déroulante
        self.lstAction.Clear()
        self.lstAction.AppendItems(self.search_identifiant_action())
        self.lstAction.SetSelection(0)

        self.Centre()
        self.CreateStatusBar()
        self.SetStatusText("Créé par Philippe CLARET :-)")
        self.Bind(wx.EVT_BUTTON, self.onBoutonClick, self.btnEcrire)
        self.Show(True)               

    def onBoutonClick(self, event):
        # on appelle une méthode pour rechercher toutes les valeurs de la sélection
        rows = self.search_price_by_action(self.lstAction.GetStringSelection())
        # on appelle une méthode pour écrire le résultat dans un fichier Excel
        self.write_in_file(self.lstAction.GetStringSelection(), rows)      
        
        # Ajouter un messaqge pour spécifier le répertoire de travail actuel
        cwd = os.getcwd()
        dlg = MyMessageDialog(self, cwd, "Répertoire du fichier " + self.lstAction.GetStringSelection())
        dlg.ShowModal()
        
        # Fermer la fenêtre principale
        self.Close()
        
class MyMessageDialog(wx.MessageDialog):
    def __init__(self, parent, message, title, style=wx.OK, pos=wx.DefaultPosition) :
        wx.MessageDialog.__init__(self, parent, message, title, style=wx.OK, pos=wx.DefaultPosition)  

class Interface(wx.App):   # Application et fenêtre
    def __init__(self):
        wx.App.__init__(self)   
        self.frame = MyFrame("Historique d'une action boursière")
        self.SetTopWindow(self.frame)
        self.go()
    def go(self):          # Exécution de l'application.
        self.frame.Show(True)        
        self.MainLoop()

if __name__ == '__main__':
    #app = wx.App()
    #frame = MyFrame("Historique d'une action boursière")
    #app.MainLoop()
    
    # Création d'une interface pour installer le gestionnaire et ouvrir la fenêtre principale
    app=Interface()
