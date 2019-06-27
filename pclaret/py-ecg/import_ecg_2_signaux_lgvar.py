### IMPORT ECG FILE

""" Pre requis : les fichiers .dat et .hea doivent avoir le meme nom """

import wx
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment

class MyFrame(wx.Frame):
   
    # Importation des ordonnées du format binaire au format entier dans une liste
    def import_data(self):
    
        path = self.txtAction1.GetValue()
        os.chdir(path)     
        
        nom_fic = self.lstAction2.GetStringSelection()        
         
        # import & header
        # Charger les données d'un fichier texte (première ligne, colonne 1 à 3) et mettre dans un tableau
        hea0 = np.genfromtxt(nom_fic + '.hea', dtype = int, delimiter=' ', usecols=(1, 2, 3))[0] 
           
        # Joindre le tableau avec un tableau de 3 zéros (concaténation)
        hea0 = np.hstack((hea0, np.zeros(3, dtype= int)))
     
         # Charger les données d'un fichier texte (2ième et 3ième ligne, colonne 1 à 6) et mettre dans un tableau
        hea1 = np.genfromtxt(nom_fic + '.hea', dtype = int, delimiter=' ', skip_header = 1, usecols=(1, 2, 3, 4, 5, 6))
         
        # Empiler les 2 tableaux verticalement en séquence (par rangée)
        hea = np.vstack((hea0, hea1))
      
        nb_sig = hea[0, 0]      
        #sam_freq = hea[0, 1]  
        length = hea[0, 2]  
    
        # check first / checksum
        ckf = hea[1:, 4]    
        cks = hea[1:, 5]      
        
        # import 212
        sto = np.zeros((nb_sig, length), dtype=np.int16)
                
        n = 12    
        # complément à 2 sur n bits en binaire
        def c2(n):
            return lambda B : B if (B < 2**(n-1) ) else B - 2**n
    
        c12 = c2(n)   
               
        # Ouverture du fichier de données .dat
        fich = open(nom_fic + ".dat", "rb")
    
        # Lecture du fichier de données au format .dat avec 2 signaux
        byte = fich.read(1)     
        ind = 0;
        while byte: 
            l = int.from_bytes(byte, 'big')         
        
            byte = fich.read(1)                   
            m = int.from_bytes(byte, 'big')      
        
            byte = fich.read(1)                    
            r = int.from_bytes(byte, 'big')       
        
            # Ordonnées du premier signal
            m0 = ( m & int('0x0f', 16) ) << 8    
            sto[0, ind] = c12( l + m0 )          
        
            # Ordonnées du deuxième signal
            m1 = ( m & int('0xf0', 16) ) << 4    
            sto[1, ind] = c12 ( m1 + r )         
            
            byte = fich.read(1)
            ind +=1
        fich.close()  
    
        print("")           
        
        check = True
        
        # check first and checksum
        if ( np.prod(sto[:, 0]==ckf) ):    
            print('import initial ok')
        else:
            print('import initial ko') 
            print('sto[:, 0] %s' %sto[:, 0])
            print('ckf %s' %ckf)
            dlg = wx.MessageDialog(self, "Les premières mesures ne sont pas correctes. Veuillez changer le fichier de données.", "Contrôle de cohérence des données", style=wx.OK , pos=wx.DefaultPosition)
            dlg.ShowModal() 
            self.txtLabel1.SetLabel("Nombre total de points sur le signal : 0")
            self.lstAction3.Clear()  
            self.lstAction3.Enable(False)
            self.lstAction4.Clear()  
            self.lstAction4.Enable(False)
            self.lstAction5.Clear()  
            self.lstAction5.Enable(False)
            self.btnEcrire.Enable(False) 
            check = False
          
        if ( np.prod( np.int16( np.sum(sto, axis = 1) ) == cks ) ):    
            print('checksum ok')              
        else:
            print('checksum ko')
            print ('sto %s' %np.int16( np.sum(sto, axis = 1)))
            print ('cks %s' %cks) 
            dlg = wx.MessageDialog(self, "La somme des mesures n'est pas correcte. Veuillez changer le fichier de données.", "Contrôle de cohérence des données", style=wx.OK , pos=wx.DefaultPosition)
            dlg.ShowModal() 
            self.txtLabel1.SetLabel("Nombre total de points sur le signal : 0")
            self.lstAction3.Clear()  
            self.lstAction3.Enable(False)
            self.lstAction4.Clear()  
            self.lstAction4.Enable(False)
            self.lstAction5.Clear()  
            self.lstAction5.Enable(False)
            self.btnEcrire.Enable(False) 
            check = False
            
        return sto, length, nb_sig, check
    
    # Découpage des données sto en n fichiers de p points de mesure
    def cut_sto (self, sto, length, nb_sig):
        
        nbr_pts = int(self.lstAction3.GetStringSelection())  
        nbr_fic = int(self.lstAction4.GetStringSelection())
                          
        i = 1
        ind = 0
        while i <= nbr_fic: 
            j = 0
            stw = np.zeros((nb_sig, nbr_pts), dtype=np.int16)
            while j <= nbr_pts - 1 and ind <= length - 1: 
                stw[0, j] = sto[0, ind]
                stw[1, j] = sto[1, ind]
                j += 1
                ind += 1
            self.write_in_file(stw, i, j, length)   
            i += 1
                  
    # Ecriture des données stw dans un fihcier EXCEL
    def write_in_file(self, stw, num_fic, nbr_pts, length):
         
        nom_fic = self.lstAction2.GetStringSelection()
        nom_sig = self.lstAction5.GetStringSelection()
        ind_sig = self.lstAction5.GetSelection()   
                   
        # **************************************************************************************
        # Ecrire du résultat dans un fichier Excel
        # **************************************************************************************
        # Créer un classeur
        wb = Workbook()
        # Créer une feuille
        ws = wb.active
        ws.title = "Signal " + nom_sig

        # Creér un style nommé pour les titres de colonne dans le fichier
        StyleLabel = NamedStyle(name="StyleLabel")
        StyleLabel.font = Font(bold=True, color="FFFFFF")
        StyleLabel.alignment = Alignment(horizontal='center', vertical='center')
        StyleLabel.fill = PatternFill(fgColor="318CE7", fill_type="solid")
        StyleLabel.border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   top=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))

        # Creér un style nommé pour les valeurs de colonne dans le fichier
        StyleValue = NamedStyle(name="StyleValue")
        StyleValue.alignment = Alignment(horizontal='center', vertical='center')
        StyleValue.border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'))

        # Définir la largeur des colonnes du fichier
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15            

        # Ecrire les titres de colonnes dans le fichier
        ws.cell(row=1, column=1, value="Abscisses").style = StyleLabel
        ws.cell(row=1, column=2, value="Ordonnées").style = StyleLabel              

        # Calcul de toutes les abscisses en fonction de la longueur du signal
        abs = np.arange(nbr_pts)
        
        # Remplir le fichier avec toutes les valeurs des colonnes        
        i = 2 
        j = 0
        k = 0         
        while k < nbr_pts:  
            ws.cell(row=i, column=1, value=abs[k]).style = StyleValue
            ws.cell(row=i, column=2, value=stw[ind_sig, k]).style = StyleValue
            if j == 1:
                ws.cell(row=i, column=1).fill = PatternFill(fgColor="DFF2FF", fill_type="solid")
                ws.cell(row=i, column=2).fill = PatternFill(fgColor="DFF2FF", fill_type="solid")
                j = 0
            else:
                j = 1
            i +=1
            k +=1
                 
        # Mettre une bordure inférieure sur la dernière ligne
        ws.cell(row=i-1, column=1).border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))
        ws.cell(row=i-1, column=2).border = Border(left=Side(border_style="thin", color='FF000000'),
                                   right=Side(border_style="thin", color='FF000000'),
                                   bottom=Side(border_style="thin", color='FF000000'))
                
        if nbr_pts == length:
            wb.save(nom_fic + "_" + nom_sig + ".xlsx")
        else:
            wb.save(nom_fic + "_" + nom_sig + "_" + str(num_fic) + ".xlsx")
               
        return True
    
    # Construction de la fenêtre principale    
    def __init__(self, titre):
        
        wx.Frame.__init__(self, None, 1, title=titre)
        
         # Définir les différentes zones dans la fenêtre
        frameSizer = wx.BoxSizer(wx.HORIZONTAL)
        panelSizer = wx.BoxSizer(wx.VERTICAL)
        conteneur = wx.Panel(self, 1)
    
        # Ajouter des textes dans le conteneur
        textAction1 = wx.StaticText(conteneur, 1, "Coller le chemin absolu du fichier sans son nom", style=wx.ALIGN_CENTRE)
        textAction2 = wx.StaticText(conteneur, 1, "Sélectionner un fichier", style=wx.ALIGN_CENTRE)
        textAction3 = wx.StaticText(conteneur, 1, "Sélectionner le nombre de points souhaité dans un fichier EXCEL", style=wx.ALIGN_CENTRE)
        textAction4 = wx.StaticText(conteneur, 1, "Sélectionner le nombre de fichiers EXCEL à écrire", style=wx.ALIGN_CENTRE)
        textAction5 = wx.StaticText(conteneur, 1, "Sélectionner un signal ECG", style=wx.ALIGN_CENTRE)
        textbtnEcrire = wx.StaticText(conteneur, 1, "Ecrire les coordonnées du signal ECG", style=wx.ALIGN_CENTRE)
            
        # Ajouter une zone de texte pour récupérer le répertoire de travail
        self.txtAction1 = wx.TextCtrl(conteneur, -1, size = (300,10))
        
         # Ajouter une liste deroulante pour lister les noms de fichier 
        self.lstAction2 = wx.Choice(conteneur)
        self.lstAction2.Enable(False)   
        
        # Ajoueter un libellé pour afficher la longueur du fichier
        self.txtLabel1 = wx.StaticText(conteneur, 1, "Nombre total de points sur le signal : 0", style=wx.ALIGN_CENTRE)
       
        # Ajouter une liste deroulante pour lister le nombre de points à traiter
        self.lstAction3 = wx.Choice(conteneur)
        self.lstAction3.Enable(False)
        
        # Ajouter une liste deroulante pour lister le nombre de fichier à écrire
        self.lstAction4 = wx.Choice(conteneur)
        self.lstAction4.Enable(False)
             
        # Ajouter une liste deroulante pour lister les signaux présents dans le fichier
        self.lstAction5 = wx.Choice(conteneur)
        self.lstAction5.Enable(False)
        
        # Ajouter un bouton pour écrire les coordonnées des signaux dans un fichier Excel
        self.btnEcrire = wx.Button(conteneur, -1, "Exécuter")
        self.btnEcrire.Enable(False)
            
        # Ajouter verticalement tous les contrôles
        panelSizer.AddSpacer(20)
        panelSizer.Add(textAction1, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.txtAction1, 1, wx.ALIGN_CENTRE)
        panelSizer.AddSpacer(20)
        panelSizer.Add(textAction2, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.lstAction2, 1, wx.ALIGN_CENTRE) 
        panelSizer.AddSpacer(20)        
        panelSizer.Add(self.txtLabel1, 1, wx.ALIGN_CENTRE)        
        panelSizer.AddSpacer(20)        
        panelSizer.Add(textAction3, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.lstAction3, 1, wx.ALIGN_CENTRE)
        panelSizer.AddSpacer(20)
        panelSizer.Add(textAction4, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.lstAction4, 1, wx.ALIGN_CENTRE)        
        panelSizer.AddSpacer(20)
        panelSizer.Add(textAction5, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.lstAction5, 1, wx.ALIGN_CENTRE)        
        panelSizer.AddSpacer(20)
        panelSizer.Add(textbtnEcrire, 1, wx.ALIGN_CENTRE)
        panelSizer.Add(self.btnEcrire, 1, wx.ALIGN_CENTRE)
        panelSizer.AddSpacer(20)
    
        conteneur.SetSizer(panelSizer)
        frameSizer.Add(conteneur, 1, wx.EXPAND)
        self.SetSizer(frameSizer)
        frameSizer.SetSizeHints(self)
        self.SetSize((550, 350))
           
        self.Centre()
        self.CreateStatusBar()
        self.SetStatusText("Créé par Philippe CLARET :-)")        
        self.Bind(wx.EVT_TEXT, self.onTextChanged, self.txtAction1)
        self.Bind(wx.EVT_CHOICE, self.onSelectedItem2, self.lstAction2) 
        self.Bind(wx.EVT_CHOICE, self.onSelectedItem3, self.lstAction3)
        self.Bind(wx.EVT_BUTTON, self.onBoutonClick, self.btnEcrire)
        self.Show(True)

    # Remplissage des différentes listes déroulantse 
    def onTextChanged(self, event):
                    
        # liste déroulante des fichiers présents dans la base de données
        lstNom = self.search_lstAction2(self.txtAction1.GetValue())
                
        if lstNom != []:    
            self.lstAction2.Clear()        
            self.lstAction2.AppendItems(lstNom)
            self.lstAction2.SetSelection(0)  
            self.lstAction2.Enable(True)      
            
            # Mise à jour des critères en fonction du fichier sélectionné
            self.onSelectedItem2(event)           
        else:
            self.lstAction2.Clear()        
            self.lstAction2.Enable(False)
            self.txtLabel1.SetLabel("Nombre total de points sur le signal : 0")
            self.lstAction3.Clear()  
            self.lstAction3.Enable(False)
            self.lstAction4.Clear()  
            self.lstAction4.Enable(False)
            self.lstAction5.Clear()  
            self.lstAction5.Enable(False)
            self.btnEcrire.Enable(False)             
             
    # Mise à jour jour des critères en fonction du fichier sélectionné
    def onSelectedItem2(self, event):
        
        os.chdir(self.txtAction1.GetValue()) 
        nom_fic = self.lstAction2.GetStringSelection()
        
        len_sig = self.search_length_enreg(nom_fic)
                    
        # Mise à jour de la longueur du signal pour le premier fichier
        self.txtLabel1.SetLabel("Nombre total de points sur le signal : " + str(len_sig))
       
        # Liste déroulante affichant le nombre de points à traiter
        self.lstAction3.Clear() 
        self.lstAction3.AppendItems("1000")
        self.lstAction3.AppendItems("2000")
        self.lstAction3.AppendItems("5000")
        self.lstAction3.AppendItems("10000")
        self.lstAction3.AppendItems("15000")
        self.lstAction3.AppendItems("20000")
        self.lstAction3.AppendItems(str(len_sig))
        self.lstAction3.SetSelection(0)  
        self.lstAction3.Enable(True)
        
        # liste déroulante affichant le nombre de fichiers EXCEL à écrire
        Nbr_fic = self.search_lstNbr_fic(len_sig, self.lstAction3.GetStringSelection())
        self.lstAction4.Clear()        
        self.lstAction4.AppendItems(Nbr_fic)
        self.lstAction4.SetSelection(0)  
        self.lstAction4.Enable(True)
           
        try:                      
            # Noms des signaux
            lstsig = np.genfromtxt(nom_fic + '.hea', dtype = str, delimiter=' ', skip_header = 1, usecols=(8))
             
            # Remplissage de la liste déroulante des signaux présents dans le fichier sélectionné
            self.lstAction5.Clear()
            self.lstAction5.AppendItems(lstsig)
            self.lstAction5.SetSelection(0)  
        except ValueError: 
            dlg = wx.MessageDialog(self, "Nom des signaux ECG inconnus", "Absence de données détectée", style=wx.OK , pos=wx.DefaultPosition)
            dlg.ShowModal() 
            
            # Remplissage de la liste déroulante des signaux présents dans le fichier sélectionné
            self.lstAction5.Clear()
            self.lstAction5.AppendItems("Signal 1")
            self.lstAction5.AppendItems("Signal 2")
            self.lstAction5.SetSelection(0)  
            
        self.lstAction5.Enable(True)
                 
        # Activation du bouton d'action       
        self.btnEcrire.Enable(True)
            
    # Mise à jour jour suite au changement du nombre de points à traiter
    def onSelectedItem3(self, event):
        
        # liste déroulante affichant le nombre de fichiers EXCEL à écrire
        lstNom = self.search_lstAction2(self.txtAction1.GetValue())
        len_sig = self.search_length_enreg(lstNom[0])
        Nbr_fic = self.search_lstNbr_fic(len_sig, self.lstAction3.GetStringSelection())
        
        # Remplissage de la liste déroulante du nombre de fichiers EXCEL à écrire
        self.lstAction4.Clear()
        self.lstAction4.AppendItems(Nbr_fic)
        self.lstAction4.SetSelection(0)  
        
    # Récupération de la longueur d'un fichier de données    
    def search_length_enreg(self, nom_fic):
        
         os.chdir(self.txtAction1.GetValue()) 
         length = np.genfromtxt(nom_fic + '.hea', dtype = int, delimiter=' ', usecols=(3))[0]
         
         return length
   
    # Ecriture dans le fichier EXCEL puis affichage d'un message de confirmation
    def onBoutonClick(self, event):
          
        # Récupération des données dans le fichier .dat
        sto, length, nb_sig, check = self.import_data()
        
        # Ecriture des données dans un fichier Excel
        if check:
            self.cut_sto (sto, length, nb_sig)          
            
            # Ajouter un message pour spécifier le répertoire de travail actuel
            cwd = os.getcwd()        
            dlg = wx.MessageDialog(self, cwd, "Répertoire du fichier " + self.lstAction2.GetStringSelection() + ".xlsx", style=wx.OK , pos=wx.DefaultPosition)
            dlg.ShowModal()
            
            # Fermer la fenêtre principale
            self.Close()         
        
    # Récupération de tous les noms de fichiers présents dans la base de données 
    def search_lstAction2(self, nom_rep):
        
        lstNom = [] 
        
        try:             
            lstFic = os.listdir(nom_rep) 
            for fic in lstFic:             
                if fic[len(fic)-3:] == "hea":
                    lstNom.append(fic[:len(fic)-4])
        except FileNotFoundError:            
            dlg = wx.MessageDialog(self, "Répertoire non trouvé. Veuillez vérifier et saisir à nouveau !", "Erreur détectée", style=wx.OK , pos=wx.DefaultPosition)
            dlg.ShowModal()           
            
        return lstNom
                     
    # Récupération des nombres de fichiers EXCEL à écrire 
    def search_lstNbr_fic(self, len_sig, nbr_pts):
        
        nbr_fic = [] 
        
        a = int(len_sig) / int(nbr_pts)
        b = int(len_sig) % int(nbr_pts)              
        
        i = 1       
        while i <= int(a):
             nbr_fic.append(str(i))   
             i += 1
                              
        if b > 0:
            nbr_fic.append(str(i)) 
        
        return nbr_fic 
      
class MyGauge(wx.Frame):
    
   def __init__(self, parent, id, title):
       Frame.__init__(self, parent, id, title)
       self.timer = wx.Timer(self, 1)
	   self.count = 0
	   self.Bind(wx.EVT_TIMER, self.OnTimer, self.timer)
	   panel = wx.Panel(self, -1)
        # Toutes les 100ms, on incrémente la barre de progression qui possede 50 pas d'incrémentations.
		self.gauge = wx.Gauge(panel, -1, 50, size=(250, 25))
		self.timer.Start(100)
        
    def OnTimer(self, event):
		self.count = self.count +1
		self.gauge.SetValue(self.count)
        
class Interface(wx.App):   # Application et fenêtre
    
    def __init__(self):
        wx.App.__init__(self)   
        self.frame = MyFrame("Coordonnées des signaux d'électrocardiogramme")
        self.SetTopWindow(self.frame)
        self.go()
    def go(self):          # Exécution de l'application.
        self.frame.Show(True)        
        self.MainLoop()

if __name__ == '__main__':
    #app = wx.App()
    #frame = MyFrame("Coordonnées des signaux d'électrocardiogramme")
    #app.MainLoop()
    
    # Création d'une interface pour installer le gestionnaire et ouvrir la fenêtre principale
    app=Interface()
    
    
    
    
    
    